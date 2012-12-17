# coding=utf-8
"""
 Copyright (C) 2012 Jürgen Baumeister

 This program is free software: you can redistribute it and/or modify
 it under the terms of the GNU Lesser General Public License as published by
 the Free Software Foundation, either version 3 of the License, or
 (at your option) any later version.

 This program is distributed in the hope that it will be useful,
 but WITHOUT ANY WARRANTY; without even the implied warranty of
 MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
 GNU Lesser General Public License for more details.

 You should have received a copy of the GNU Lesser General Public License
 along with this program.  If not, see <http://www.gnu.org/licenses/>.
 
DS_xlsx.py
#=====================================================================================================
A data source class to extend DS_basic
This extension class uses the DS_basic data source class to model a data source, that is
based on a xlsx Excel file
"""
#==============================================================================================================
# Import section
#==============================================================================================================
from DS_basic import DS_basic
from openpyxl import load_workbook

#==============================================================================================================
# class DS_csv
#==============================================================================================================
class DS_xlsx(DS_basic):
    """Implements a data source, that uses a xlsx file to create
    the stream of records
    """

    #==============================================================================================================
    # The modified __init__()  method
    #==============================================================================================================
    def __init__(self, name='not named'):
        DS_basic.__init__(self, name)

        self.ds_file = None                 #a file object, where the data source gets the data
        self.ds_filename = None             #the file name string of the underlying data file
        self.ds_table = None                #the table object, where the data source gets the data from
        self.ds_tablename = None            #the table name, that names the table in the underlying data file
        self.ds_reader = None               #a reader object, needed for providing the data one by one
                                            # is created internally

    #==============================================================================================================
    # Methods that are specific for the derived class DS_csv
    #==============================================================================================================
    def get_filename(self):
        """Returns the filename of the Excel file, used as data source
        """
        return self.ds_filename
        
    def set_filename(self, name):
        """Set the filename for the Excel file, used as data source
        """
        self.ds_filename = name
        
    def get_tablename(self):
        """Returns the table name, that is used as data source
        """
        return self.ds_table
        
    def set_tablename(self, name):
        """Set the table name, that is used as data source
        """
        self.ds_table = name

    
    #==============================================================================================================
    # Methods that are meant to be overwritten, when a special data source is implemented
    #==============================================================================================================
    def _open_data_source(self, *args):
        """Overrides method from class data_source
        
        Opens a given file for data input
        """
        if len(args) != 0:
            # For first call to open (open())
            self.ds_filename = args[0]
            self.ds_tablename = args[1]
            self.ds_file = load_workbook(filename = args[0], use_iterators = True)
            self.ds_table = self.ds_file.get_sheet_by_name(name = args[1])
        else:
            # For reopening the file (reset())
            self.ds_file = load_workbook(filename = self.ds_filename, use_iterators = True)
            self.ds_table = self.ds_file.get_sheet_by_name(name = self.ds_tablename)
        # In any case we need a reader object to iterate over the table content 
        self.ds_reader = self.ds_table.iter_rows()

    def _close_data_source(self):
        """Overrides method from class data_source
        
        Closes the input file of the object
        """
        # As the library openpyxl does not provide explicit close() methods. we
        # just set all internal variables to None
        self.ds_reader = None
        self.ds_table = None
        self.ds_file = None
        
    def _get_next_record(self):
        """Overrides method form class data_source
        
        Gets the next line from the underlying file
        """
        # Read next record from xlsx file
        row = self.ds_reader.next()
        if len(row) == 0:
            raise StopIteration
                
        # build record
        record = {'__row':row[0].row}
        for cell in row:
            record[str(cell.column)] = cell.internal_value
            #if cell.internal_value != None:
            #   self.ds_processedBytes += len(cell.internal_value)
            
        # finished
        return record
        




